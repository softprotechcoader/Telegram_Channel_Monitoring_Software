import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def generate_empty_excel(file_path):
    """Generate an empty Excel file with headers for Phone Number, API_ID, and API_HASH."""
    wb = Workbook()
    ws = wb.active

    # Add headers with formatting
    headers = ["Phone Number", "API_ID", "API_HASH"]
    header_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    header_font = Font(color="FFFFFF")

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Save the file
    wb.save(file_path)

def load_excel(file_path):
    """Load an Excel file and validate its structure."""
    try:
        df = pd.read_excel(file_path)
        required_columns = ["Phone Number", "API_ID", "API_HASH"]
        if not all(col in df.columns for col in required_columns):
            raise ValueError(f"Excel file must contain the following columns: {', '.join(required_columns)}")
        return df
    except Exception as e:
        raise ValueError(f"Error loading Excel file: {e}")